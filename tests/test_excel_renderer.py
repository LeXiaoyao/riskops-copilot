from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from riskops.engines.report import write_business_report_excel

ROOT = Path(__file__).resolve().parents[1]
M3_SUMMARY = ROOT / "outputs" / "m3" / "m3_summary.json"
ROI_RESULTS = ROOT / "outputs" / "model_lab" / "roi_results.json"
CLI = ROOT / "scripts" / "riskops_cli.py"


def test_write_business_report_excel_generates_expected_workbook(tmp_path: Path) -> None:
    output_path = tmp_path / "m4_business_report.xlsx"

    write_business_report_excel(M3_SUMMARY, output_path, ROI_RESULTS)

    assert output_path.exists()
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["概览", "异常信号", "归因 Top5", "策略ROI"]
    assert workbook["异常信号"].max_row > 1
    assert workbook["异常信号"]["A1"].font.bold is True
    assert workbook["异常信号"]["A1"].fill.fgColor.rgb == "001F4E79"
    assert workbook["异常信号"]["A1"].font.color.rgb == "00FFFFFF"


def test_excel_anomaly_row_count_matches_m3_summary(tmp_path: Path) -> None:
    output_path = tmp_path / "m4_business_report.xlsx"
    summary = json.loads(M3_SUMMARY.read_text(encoding="utf-8"))
    expected_rows = len(summary["high_priority_anomalies"])

    write_business_report_excel(M3_SUMMARY, output_path, ROI_RESULTS)

    workbook = load_workbook(output_path)
    assert workbook["异常信号"].max_row - 1 == expected_rows


def test_render_excel_cli_can_run(tmp_path: Path) -> None:
    output_path = tmp_path / "m4_business_report.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            str(CLI),
            "render-excel",
            "--output",
            str(output_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert output_path.exists()
    assert "business report excel" in result.stdout
