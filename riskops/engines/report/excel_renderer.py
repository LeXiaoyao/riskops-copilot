"""Excel renderer for the M4 business report."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from riskops.engines.report.business_report import BusinessReportInputError, load_m3_summary

HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DATA_ROOT = PROJECT_ROOT / "synthetic_data"
CASE_DETAIL_HEADERS = [
    "case_id",
    "vendor_id",
    "region",
    "dpd_bucket",
    "balance_segment",
    "last_action_type",
    "last_outcome",
    "outstanding_amount",
]


class ExcelReportInputError(BusinessReportInputError):
    """Raised when Excel report inputs are missing or malformed."""


def write_business_report_excel(
    m3_summary_path: Path,
    output_path: Path,
    roi_results_path: Path,
    data_root: Path = DATA_ROOT,
) -> dict[str, Any]:
    summary = load_m3_summary(m3_summary_path)
    roi_results = load_roi_results(roi_results_path)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "概览"

    _write_overview_sheet(overview, summary)
    _write_table(
        workbook.create_sheet("异常信号"),
        ["metric_code", "metric_name_cn", "current_value", "baseline_value", "relative_change", "severity"],
        _anomaly_rows(summary),
    )
    _write_table(
        workbook.create_sheet("归因 Top5"),
        ["dimension_name", "dimension_value", "contribution_score", "recommended_action"],
        _driver_rows(summary),
    )
    _write_table(
        workbook.create_sheet("策略ROI"),
        ["scenario_id", "roi_ratio", "estimated_cost", "estimated_benefit", "payback_months"],
        _roi_rows(roi_results),
    )
    _write_table(
        workbook.create_sheet("明细_案件"),
        CASE_DETAIL_HEADERS,
        _case_detail_rows(data_root),
    )
    _write_vendor_region_pivot(workbook.create_sheet("维度透视_供应商×线路"), _vendor_region_pivot_rows(data_root))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"output_path": str(output_path), "anomaly_count": len(_anomaly_rows(summary))}


def load_roi_results(input_path: Path) -> dict[str, Any]:
    if not input_path.exists():
        raise ExcelReportInputError(f"ROI 输入文件不存在：{input_path}。请先运行 render-model-lab。")
    try:
        payload = json.loads(input_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ExcelReportInputError(f"ROI 输入文件不是合法 JSON：{input_path}。错误：{exc}") from exc
    if not isinstance(payload, dict):
        raise ExcelReportInputError(f"ROI 输入文件结构不符合预期：{input_path} 顶层必须是 JSON object。")
    return payload


def _write_overview_sheet(sheet: Worksheet, summary: dict[str, Any]) -> None:
    overview = _as_dict(summary.get("anomaly_overview"))
    attribution = _as_dict(summary.get("m1_d7_attribution_summary"))
    target = _as_dict(attribution.get("target_anomaly")) or _as_dict(summary.get("attribution_target_anomaly"))
    target_metric_code = attribution.get("target_metric_code") or target.get("metric_code")
    target_metric_name = attribution.get("target_metric_name_cn") or target.get("metric_name_cn")

    rows = [
        ("anomaly_count", overview.get("anomaly_count")),
        ("severity_counts", _format_severity_counts(_as_dict(overview.get("severity_counts")))),
        ("target_metric", _format_target_metric(target_metric_code, target_metric_name)),
        ("relative_change", target.get("relative_change")),
        ("生成时间", datetime.now(UTC).replace(microsecond=0).isoformat()),
    ]
    _write_table(sheet, ["field", "value"], rows)


def _write_table(sheet: Worksheet, headers: list[str], rows: list[tuple[Any, ...]]) -> None:
    sheet.append(headers)
    _style_header(sheet)
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    sheet.freeze_panes = "A2"
    _autosize_columns(sheet)


def _write_vendor_region_pivot(sheet: Worksheet, rows: list[dict[str, Any]]) -> None:
    regions = sorted({str(row["region"]) for row in rows if row.get("region")})
    vendors = sorted({str(row["vendor_id"]) for row in rows if row.get("vendor_id")})
    recovery_by_cell = {(str(row["vendor_id"]), str(row["region"])): row["recovery_rate"] for row in rows}

    sheet.append(["vendor_id / region", *regions, "合计"])
    _style_header(sheet)

    total_column = len(regions) + 2
    for row_index, vendor_id in enumerate(vendors, start=2):
        values = [_excel_value(recovery_by_cell.get((vendor_id, region))) for region in regions]
        sheet.append([vendor_id, *values, f"=SUM(B{row_index}:{get_column_letter(total_column - 1)}{row_index})"])

    total_row = len(vendors) + 2
    sheet.cell(row=total_row, column=1, value="合计")
    for column_index in range(2, total_column):
        column_letter = get_column_letter(column_index)
        sheet.cell(row=total_row, column=column_index, value=f"=SUM({column_letter}2:{column_letter}{total_row - 1})")
    total_column_letter = get_column_letter(total_column)
    sheet.cell(row=total_row, column=total_column, value=f"=SUM({total_column_letter}2:{total_column_letter}{total_row - 1})")

    sheet.freeze_panes = "B2"
    _autosize_columns(sheet)


def _style_header(sheet: Worksheet) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(bold=True, color=HEADER_FONT)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        width = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            width = max(width, len(str(cell.value)) if cell.value is not None else 0)
        sheet.column_dimensions[column_letter].width = min(max(width + 2, 12), 48)


def _anomaly_rows(summary: dict[str, Any]) -> list[tuple[Any, ...]]:
    anomalies = [item for item in _as_list(summary.get("high_priority_anomalies")) if isinstance(item, dict)]
    return [
        (
            anomaly.get("metric_code"),
            anomaly.get("metric_name_cn"),
            anomaly.get("recent_value"),
            anomaly.get("baseline_value"),
            anomaly.get("relative_change"),
            anomaly.get("severity"),
        )
        for anomaly in anomalies
    ]


def _driver_rows(summary: dict[str, Any]) -> list[tuple[Any, ...]]:
    attribution = _as_dict(summary.get("m1_d7_attribution_summary"))
    drivers = [item for item in _as_list(attribution.get("top_drivers")) if isinstance(item, dict)][:5]
    return [
        (
            driver.get("dimension_name"),
            driver.get("dimension_value"),
            driver.get("contribution_score"),
            driver.get("recommended_action"),
        )
        for driver in drivers
    ]


def _roi_rows(roi_results: dict[str, Any]) -> list[tuple[Any, ...]]:
    results = [item for item in _as_list(roi_results.get("results")) if isinstance(item, dict)]
    return [
        (
            result.get("scenario_id"),
            result.get("roi_ratio"),
            result.get("action_cost"),
            result.get("gross_benefit"),
            _payback_months(result.get("payback_period_days")),
        )
        for result in results
    ]


def _case_detail_rows(data_root: Path, sample_size: int = 200) -> list[tuple[Any, ...]]:
    case_status = _latest_partition(_read_parquet(data_root, "dws/dws_case_status_snapshot_di.parquet"))
    case_status = case_status.sort_values("case_id").head(sample_size).copy()

    case_dim = _read_parquet(data_root, "dim/dim_case.parquet")
    case_dim = case_dim[["case_id", "balance_segment"]].drop_duplicates("case_id")
    case_status = case_status.merge(case_dim, on="case_id", how="left")

    capacity = _latest_partition(_read_parquet(data_root, "dws/dws_vendor_line_capacity_di.parquet"))
    capacity = capacity[["vendor_id", "line_id", "region"]].drop_duplicates(["vendor_id", "line_id"])
    case_status = case_status.merge(capacity, on=["vendor_id", "line_id"], how="left")

    process = _latest_partition(_read_parquet(data_root, "dws/dws_collection_process_wide_di.parquet"))
    process = process[
        ["case_id", "action_count", "ai_action_count", "connected_count", "ptp_count", "ptp_fulfilled_count"]
    ].drop_duplicates("case_id")
    case_status = case_status.merge(process, on="case_id", how="left", suffixes=("", "_process"))

    case_status["last_action_type"] = case_status.apply(_last_action_type, axis=1)
    case_status["last_outcome"] = case_status.apply(_last_outcome, axis=1)

    return [
        tuple(row.get(column) for column in CASE_DETAIL_HEADERS)
        for row in case_status[CASE_DETAIL_HEADERS].to_dict("records")
    ]


def _vendor_region_pivot_rows(data_root: Path) -> list[dict[str, Any]]:
    vendor = _read_parquet(data_root, "ads/ads_vendor_performance_di.parquet")
    vendor = (
        vendor.groupby("vendor_id", as_index=False)
        .agg(ptp_rate=("ptp_rate", "mean"), ptp_keep_rate=("ptp_keep_rate", "mean"))
        .assign(recovery_rate=lambda frame: frame["ptp_rate"] * frame["ptp_keep_rate"])
    )

    capacity = _latest_partition(_read_parquet(data_root, "dws/dws_vendor_line_capacity_di.parquet"))
    vendor_region = capacity[["vendor_id", "region"]].drop_duplicates(["vendor_id", "region"])
    rows = vendor_region.merge(vendor[["vendor_id", "recovery_rate"]], on="vendor_id", how="left")
    rows = rows.sort_values(["vendor_id", "region"])
    return rows[["vendor_id", "region", "recovery_rate"]].to_dict("records")


def _read_parquet(data_root: Path, relative_path: str) -> pd.DataFrame:
    path = data_root / relative_path
    if not path.exists():
        raise ExcelReportInputError(f"Excel 数据输入不存在：{path}")
    return pd.read_parquet(path)


def _latest_partition(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty or "stat_date" not in data.columns:
        return data.copy()
    latest_date = data["stat_date"].max()
    return data[data["stat_date"].eq(latest_date)].copy()


def _last_action_type(row: pd.Series) -> str:
    action_count = _number(row.get("action_count_process", row.get("action_count")))
    ai_action_count = _number(row.get("ai_action_count"))
    if action_count <= 0:
        return "NO_ACTION"
    if ai_action_count > 0:
        return "AI_ACTION"
    return "MANUAL_ACTION"


def _last_outcome(row: pd.Series) -> str:
    if _number(row.get("ptp_fulfilled_count")) > 0:
        return "PTP_FULFILLED"
    if _number(row.get("ptp_count")) > 0:
        return "PTP"
    if _number(row.get("connected_count_process", row.get("connected_count"))) > 0:
        return "CONNECTED"
    if _number(row.get("action_count_process", row.get("action_count"))) > 0:
        return "NOT_CONNECTED"
    return "NO_ACTION"


def _payback_months(value: Any) -> float | None:
    if not isinstance(value, int | float):
        return None
    return round(value / 30, 4)


def _format_severity_counts(severity_counts: dict[str, Any]) -> str:
    return ", ".join(f"{key}={severity_counts.get(key, 0)}" for key in ("high", "medium", "low"))


def _format_target_metric(metric_code: Any, metric_name: Any) -> str:
    if metric_code and metric_name:
        return f"{metric_name} ({metric_code})"
    return str(metric_name or metric_code or "")


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _number(value: Any) -> float:
    numeric_value = pd.to_numeric(value, errors="coerce")
    return 0.0 if pd.isna(numeric_value) else float(numeric_value)


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value
