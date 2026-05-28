"""DeepSeek-style chat TUI for RiskOps Copilot."""

from __future__ import annotations

import asyncio
import os
from pathlib import Path

from rich.align import Align
from rich.markup import escape
from textual.app import App, ComposeResult
from textual.containers import Container, VerticalScroll
from textual.message import Message
from textual.widgets import Label, Markdown, Static, TextArea

from riskops.agents.orchestrator import RiskOpsOrchestrator
from riskops.tui.context_loader import context_summary, read_output

ROOT = Path(__file__).resolve().parents[2]

WELCOME_MESSAGE = """欢迎使用 RiskOps Copilot TUI。
当前已加载：M3 异常报告 / 归因摘要 / 策略 ROI / Briefing。
直接输入问题，或使用 /help 查看命令。"""

HELP_TEXT = """可用命令：
- /help：显示所有命令列表
- /clear：清空聊天历史
- /context：显示当前加载的数据摘要
- /summary：展示 outputs/m3/m3_summary.md
- /anomaly：展示 outputs/anomalies/anomaly_results.md
- /drivers：展示 outputs/attribution/attribution_summary.md
- /roi：展示 outputs/model_lab/roi_summary.md
- /briefing：展示 outputs/copilot/briefing.md
- /model：切换模型 deepseek-chat / deepseek-reasoner
- /qc：展示 outputs/qc/qc_summary.md（新增）
- /script：展示 outputs/script/script_summary.md（新增）
- /report：异步导出 outputs/reports/ 周报产物（新增）
- /vendor：展示 outputs/vendor_review.xlsx 的概览 sheet（新增）"""

SLASH_COMMANDS = {
    "/help",
    "/clear",
    "/context",
    "/summary",
    "/anomaly",
    "/drivers",
    "/roi",
    "/briefing",
    "/model",
    "/qc",
    "/script",
    "/report",
    "/vendor",
}


def read_qc_summary(root: Path | None = None) -> str:
    return _read_markdown_output(root, Path("outputs/qc/qc_summary.md"), "质检摘要")


def read_script_summary(root: Path | None = None) -> str:
    return _read_markdown_output(root, Path("outputs/script/script_summary.md"), "话术摘要")


def read_vendor_overview(root: Path | None = None) -> str:
    base = root or ROOT
    relative_path = Path("outputs/vendor_review.xlsx")
    path = base / relative_path
    if not path.exists():
        return f"未找到 {relative_path}，请先跑 run_all.sh"

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return f"读取 {relative_path} 失败：{exc}"

    if "概览" not in workbook.sheetnames:
        workbook.close()
        return f"未找到 {relative_path} 的「概览」sheet，请先重新生成供应商绩效文件"

    sheet = workbook["概览"]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return f"{relative_path} 的「概览」sheet 为空"

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    body_rows = rows[1:]
    lines = [f"## 供应商绩效概览\n\n来源：{relative_path}"]
    for row_index, row in enumerate(body_rows, start=1):
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        lines.append(f"\n### 记录 {row_index}")
        for header, value in zip(headers, values, strict=False):
            if not header:
                continue
            display_value = "" if value is None else str(value)
            lines.append(f"- **{header}**：{display_value}")

    if len(lines) == 1:
        lines.append("\n「概览」sheet 没有可展示的数据行")
    return "\n".join(lines)


def run_report_export(root: Path | None = None) -> str:
    from riskops.engines.report import (
        BusinessReportInputError,
        write_business_report,
        write_business_report_excel,
        write_business_report_ppt,
        write_business_report_word,
    )

    base = root or ROOT
    m3_summary = base / "outputs" / "m3" / "m3_summary.json"
    roi_results = base / "outputs" / "model_lab" / "roi_results.json"
    report_dir = base / "outputs" / "reports"
    output_md = report_dir / "m4_business_report.md"
    output_html = report_dir / "m4_business_report.html"
    output_feishu = report_dir / "weekly_report.feishu.md"
    output_xlsx = report_dir / "m4_business_report.xlsx"
    output_pptx = report_dir / "m4_business_report.pptx"
    output_docx = report_dir / "m4_business_report.docx"

    try:
        write_business_report(m3_summary, output_md, output_html, output_feishu)
        write_business_report_excel(m3_summary, output_xlsx, roi_results)
        write_business_report_ppt(m3_summary, output_pptx, roi_results)
        write_business_report_word(m3_summary, output_docx, roi_results)
    except BusinessReportInputError as exc:
        return f"报告导出失败：{exc}"

    generated = sorted(path for path in report_dir.iterdir() if path.is_file())
    lines = ["报告导出完成。outputs/reports/ 当前产物："]
    lines.extend(f"- {_display_path(path, base)}" for path in generated)
    return "\n".join(lines)


def _read_markdown_output(root: Path | None, relative_path: Path, title: str) -> str:
    base = root or ROOT
    path = base / relative_path
    if not path.exists():
        return f"未找到 {relative_path}，请先跑 run_all.sh"
    return f"## {title}\n\n{path.read_text(encoding='utf-8').strip()}"


def _display_path(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


class Composer(TextArea):
    """Three-line composer that sends on Enter and inserts newline on Shift+Enter."""

    DEFAULT_CSS = """
    Composer {
        height: 3;
        border: solid #3b82f6;
        background: #0d1117;
        color: white;
    }
    """

    class Submitted(Message):
        def __init__(self, value: str) -> None:
            self.value = value
            super().__init__()

    def __init__(self) -> None:
        super().__init__(
            "",
            id="composer-input",
            placeholder="编写问题，或使用 /...",
            show_line_numbers=False,
            soft_wrap=True,
        )

    def on_key(self, event) -> None:  # type: ignore[no-untyped-def]
        if event.key == "shift+enter":
            self.insert("\n")
            event.stop()
            event.prevent_default()
            return
        if event.key == "enter":
            value = self.text.strip()
            if value:
                self.post_message(self.Submitted(value))
                self.clear()
            event.stop()
            event.prevent_default()


class RiskOpsTUIApp(App[None]):
    """RiskOps Copilot chat console backed by DeepSeek."""

    TITLE = "RiskOps Copilot TUI"
    BINDINGS = [("q", "quit", "退出")]
    ALLOWED_MODELS = ("deepseek-chat", "deepseek-reasoner")

    CSS = """
    Screen {
        background: #0d1117;
        color: white;
        layout: vertical;
    }

    #topbar {
        height: 1;
        background: #1f2937;
        color: white;
        padding: 0 1;
        layout: horizontal;
    }

    #topbar-left {
        width: 1fr;
    }

    #topbar-right {
        width: auto;
        color: #9ca3af;
    }

    #chat-history {
        height: 1fr;
        padding: 1 2;
        background: #0d1117;
    }

    .user-message {
        color: #3b82f6;
        text-align: right;
        margin: 1 0;
    }

    .ai-message {
        color: white;
        margin: 1 0;
    }

    .command-output {
        color: #22c55e;
        margin: 1 0;
    }

    .tool-call {
        color: #f59e0b;
        margin: 1 0;
    }

    .route-info {
        color: #3b82f6;
        margin: 1 0;
    }

    #composer-shell {
        height: 4;
        padding: 0 1;
        background: #0d1117;
    }

    #composer-label {
        height: 1;
        color: #9ca3af;
    }

    #footer {
        height: 1;
        background: #1f2937;
        color: #9ca3af;
        padding: 0 1;
    }
    """

    def __init__(self, *, api_key: str | None = None, model: str = "deepseek-chat") -> None:
        super().__init__()
        self.api_key = api_key or os.getenv("DEEPSEEK_API_KEY")
        self.model = model
        self.messages: list[dict[str, str]] = []

    def compose(self) -> ComposeResult:
        with Container(id="topbar"):
            yield Static(">_ RiskOps Copilot TUI", id="topbar-left")
            yield Static(self._model_status(), id="topbar-right")
        yield VerticalScroll(id="chat-history")
        with Container(id="composer-shell"):
            yield Label("Composer", id="composer-label")
            yield Composer()
        yield Static(self._footer_text(), id="footer")

    async def on_mount(self) -> None:
        await self._append_ai(WELCOME_MESSAGE)
        self.query_one("#composer-input", Composer).focus()

    async def on_composer_submitted(self, event: Composer.Submitted) -> None:
        await self._handle_input(event.value)

    async def _handle_input(self, text: str) -> None:
        await self._append_user(text)
        if text.startswith("/"):
            await self._handle_command(text)
            return
        await self._ask_deepseek(text)

    async def _handle_command(self, text: str) -> None:
        command = text.split(maxsplit=1)[0]
        if command == "/help":
            await self._append_command(HELP_TEXT)
            return
        if command == "/clear":
            await self.query_one("#chat-history", VerticalScroll).remove_children()
            await self._append_command("聊天历史已清空。")
            return
        if command == "/context":
            await self._append_command(context_summary())
            return
        if command in {"/summary", "/anomaly", "/drivers", "/roi", "/briefing"}:
            await self._append_markdown(read_output(command.removeprefix("/")))
            return
        if command == "/qc":
            await self._append_markdown(read_qc_summary())
            return
        if command == "/script":
            await self._append_markdown(read_script_summary())
            return
        if command == "/vendor":
            await self._append_markdown(read_vendor_overview())
            return
        if command == "/report":
            await self._handle_report_export()
            return
        if command == "/model":
            self._switch_model()
            await self._append_command(f"当前模型：{self.model}")
            return
        await self._append_command("未知命令。输入 /help 查看可用命令。")

    async def _handle_report_export(self) -> None:
        status = await self._append_command("导出中...")
        result = await asyncio.to_thread(run_report_export)
        status.update(escape(result))

    async def _ask_deepseek(self, text: str) -> None:
        if not self.api_key:
            await self._append_command("请先设置 export DEEPSEEK_API_KEY=your_key")
            return

        self.messages.append({"role": "user", "content": text})
        widget = await self._append_ai("")
        answer = ""
        try:
            orchestrator = RiskOpsOrchestrator(api_key=self.api_key, model=self.model)
            for event in orchestrator.run(text):
                if isinstance(event, dict) and event.get("type") == "tool_call":
                    await self._append_tool_call(event)
                    continue
                token = event if isinstance(event, str) else ""
                if token.startswith("**[→ "):
                    await self._append_route_info(token)
                    continue
                if token.lstrip().startswith("[工具调用]"):
                    await self._append_tool_call_text(token.strip())
                    continue
                answer += token
                widget.update(self._ai_renderable(answer))
                self.query_one("#chat-history", VerticalScroll).scroll_end(animate=False)
        except RuntimeError as exc:
            await self._append_command(str(exc))
            return

        self.messages.append({"role": "assistant", "content": answer})

    async def _append_user(self, text: str) -> Static:
        widget = Static(Align.right(f"[你] {escape(text)}"), classes="user-message")
        await self._mount_message(widget)
        return widget

    async def _append_ai(self, text: str) -> Static:
        widget = Static(self._ai_renderable(text), classes="ai-message")
        await self._mount_message(widget)
        return widget

    async def _append_command(self, text: str) -> Static:
        widget = Static(escape(text), classes="command-output")
        await self._mount_message(widget)
        return widget

    async def _append_tool_call(self, event: dict) -> Static:
        tool = event.get("tool", "unknown")
        params = event.get("params", {})
        row_count = event.get("row_count", 0)
        param_text = ", ".join(f"{key}={value}" for key, value in params.items())
        widget = Static(escape(f"[工具调用] {tool}({param_text}) → 返回 {row_count} 行"), classes="tool-call")
        await self._mount_message(widget)
        return widget

    async def _append_tool_call_text(self, text: str) -> Static:
        widget = Static(escape(text), classes="tool-call")
        await self._mount_message(widget)
        return widget

    async def _append_route_info(self, text: str) -> Static:
        widget = Static(escape(text.strip()), classes="route-info")
        await self._mount_message(widget)
        return widget

    async def _append_markdown(self, text: str) -> Markdown:
        widget = Markdown(text, classes="command-output")
        await self._mount_message(widget)
        return widget

    async def _mount_message(self, widget: Static | Markdown) -> None:
        history = self.query_one("#chat-history", VerticalScroll)
        await history.mount(widget)
        history.scroll_end(animate=False)

    def _ai_renderable(self, text: str) -> str:
        return f"[Copilot] {escape(text)}"

    def _switch_model(self) -> None:
        index = self.ALLOWED_MODELS.index(self.model) if self.model in self.ALLOWED_MODELS else 0
        self.model = self.ALLOWED_MODELS[(index + 1) % len(self.ALLOWED_MODELS)]
        self.query_one("#topbar-right", Static).update(self._model_status())
        self.query_one("#footer", Static).update(self._footer_text())

    def _model_status(self) -> str:
        return f"model: {self.model}  /model to switch"

    def _footer_text(self) -> str:
        return f"agent · {self.model}   q 退出   /help 帮助"


def run() -> None:
    RiskOpsTUIApp().run()


if __name__ == "__main__":
    run()
